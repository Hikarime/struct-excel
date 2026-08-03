import logging
import sys

from openpyxl.worksheet.worksheet import Worksheet
from struct_excel.database import init_db, model_to_db
from struct_excel.models import Course, Session, Enrollment, Student, Supervisor
from struct_excel.normalization import normalize_sheet
from openpyxl import Workbook, load_workbook
from struct_excel.reader import read_raw_row, read_training_list
from struct_excel.transform import (
    to_course,
    to_enrollment,
    to_session,
    to_student,
    to_supervisor,
    to_training_list,
)

logger = logging.getLogger(__name__)


def normalize_excel_sheet(
    sheet: Worksheet,
    err_path: str,
    course_sheet: Worksheet,
) -> tuple[
    list[Supervisor], list[Course], list[Student], list[Session], list[Enrollment]
]:
    err_wb = Workbook()
    err_ws = err_wb.create_sheet(sheet.title)
    normalize_sheet(sheet, err_ws)
    err_wb.save(err_path)

    normalize_sheet(course_sheet, None)

    try:
        raw_rows = read_raw_row(sheet)
    except ValueError as e:
        logger.error(str(e))
        raise

    raw_course_rows = read_training_list(course_sheet)
    training_list = to_training_list(raw_course_rows)

    supervisors = to_supervisor(raw_rows)
    courses = to_course(raw_rows, training_list)
    students = to_student(raw_rows, supervisors)
    sessions = to_session(raw_rows, courses)
    enrollments = to_enrollment(raw_rows, students, courses, sessions)

    return (supervisors, courses, students, sessions, enrollments)


def sheet_to_db(sheet: Worksheet, db_path: str, err_path: str, course_sheet: Worksheet):
    # Get entities.
    supervisors, courses, students, sessions, enrollments = normalize_excel_sheet(
        sheet,
        err_path,
        course_sheet,
    )

    # Create database.
    engine = init_db(db_path)

    # Store entities into the database.
    model_to_db(engine, supervisors)
    model_to_db(engine, courses)
    model_to_db(engine, students)
    model_to_db(engine, sessions)
    model_to_db(engine, enrollments)


def get_excel_sheet(path: str, name: str) -> Worksheet:
    wb = load_workbook(path)

    try:
        return wb[name]
    except KeyError:
        logger.error(f"invalid sheet name, available sheets: {wb.sheetnames}")
        sys.exit(1)
