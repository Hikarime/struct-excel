import streamlit as st
import sys
from io import BytesIO
from datetime import datetime, date
from pathlib import Path

import altair as alt
import pandas as pd
from openpyxl import load_workbook
from sqlmodel import Session

from struct_excel.database import init_db, model_to_db
from struct_excel.excel import normalize_excel_sheet
from struct_excel.report import ReportService

CHART_TYPES = ["Bar", "Line", "Area", "Scatter", "Pie"]


def _build_chart(
    data: dict, kind: str, show_legend: bool, show_labels: bool
) -> alt.Chart:
    df = pd.DataFrame({"name": list(data.keys()), "value": list(data.values())})
    df["label"] = df["name"] + ": " + df["value"].astype(str)
    base = alt.Chart(df)
    color = alt.Color("name:N") if show_legend else alt.Color("name:N", legend=None)
    tooltip = ["name:N", alt.Tooltip("value:Q", format="~g")]
    label = alt.Text("label:N")

    if kind == "Pie":
        chart = base.mark_arc().encode(
            theta=alt.Theta("value:Q", stack=True), color=color, tooltip=tooltip
        )
        if show_labels:
            chart += base.mark_text(fontSize=11, color="white").encode(
                theta=alt.Theta("value:Q", stack=True),
                radius=alt.value(60),
                text=label,
            )
    elif kind == "Bar":
        chart = base.mark_bar().encode(
            x="name:N", y="value:Q", color=color, tooltip=tooltip
        )
        if show_labels:
            chart += base.mark_text(dy=12, fontSize=11, color="white").encode(
                x="name:N", y="value:Q", text=label
            )
    elif kind == "Line":
        chart = base.mark_line().encode(x="name:N", y="value:Q", tooltip=tooltip)
        if show_labels:
            chart += base.mark_point(filled=True, size=50).encode(
                x="name:N", y="value:Q"
            )
            chart += base.mark_text(dy=-10, fontSize=11).encode(
                x="name:N", y="value:Q", text=label
            )
    elif kind == "Area":
        chart = base.mark_area().encode(x="name:N", y="value:Q", tooltip=tooltip)
        if show_labels:
            chart += base.mark_point(filled=True, size=50).encode(
                x="name:N", y="value:Q"
            )
            chart += base.mark_text(dy=-10, fontSize=11).encode(
                x="name:N", y="value:Q", text=label
            )
    elif kind == "Scatter":
        chart = base.mark_point(size=80, filled=True).encode(
            x="name:N", y="value:Q", color=color, tooltip=tooltip
        )
        if show_labels:
            chart += base.mark_text(dy=-10, fontSize=11).encode(
                x="name:N", y="value:Q", text=label
            )

    return chart


def _render_chart(data: dict, kind: str, show_legend: bool, show_labels: bool) -> None:
    if not data:
        st.info("No data")
        return
    st.altair_chart(_build_chart(data, kind, show_legend, show_labels), width="stretch")


def _chart_controls(key: str) -> tuple[str, bool, bool]:
    k1, c1, c2 = st.columns(3)
    kind = k1.selectbox("Chart type", CHART_TYPES, key=f"{key}_chart")
    legend_supported = kind in ("Bar", "Pie", "Scatter")
    show_legend = c1.checkbox(
        "Show legend",
        value=True,
        key=f"{key}_legend",
        disabled=not legend_supported,
        help="Legend available for Bar, Pie, Scatter" if not legend_supported else None,
    )
    show_labels = c2.checkbox("Show data labels", value=True, key=f"{key}_labels")
    return kind, show_legend, show_labels


BASE = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent.parent))
DB_FILE = BASE / "dist" / "test.db"
DB_PATH = f"sqlite:///{DB_FILE}"
ERR_XLSX = str(BASE / "dist" / "err.xlsx")

st.set_page_config(page_title="Training Report", layout="wide")
st.title("Training Report")

data_file = st.file_uploader("Registration Excel", type=["xlsx"])
course_file = st.file_uploader("Training List Excel", type=["xlsx"])

if data_file and course_file:
    wb = load_workbook(BytesIO(data_file.getvalue()))
    course_wb = load_workbook(BytesIO(course_file.getvalue()))

    c1, c2 = st.columns(2)
    sheet_name = c1.selectbox("Registration sheet", wb.sheetnames)
    course_sheet_name = c2.selectbox("Training list sheet", course_wb.sheetnames)

    if st.button("Import & Generate Report", type="primary"):
        try:
            DB_FILE.unlink(missing_ok=True)
            DB_FILE.parent.mkdir(parents=True, exist_ok=True)
            entities = normalize_excel_sheet(
                wb[sheet_name], ERR_XLSX, course_wb[course_sheet_name]
            )
            engine = init_db(DB_PATH)
            for models in entities:
                model_to_db(engine, models)

            _, _, students, sessions, enrollments = entities
            st.success(
                f"Imported {len(students)} students, {len(sessions)} sessions, "
                f"{len(enrollments)} enrollments"
            )
        except Exception as e:
            st.error(f"Import failed: {e}")

if not DB_FILE.exists():
    st.info("Upload the Excel files and click Import to generate reports.")
    st.stop()

engine = init_db(DB_PATH)

with Session(engine) as session:
    report = ReportService(session)
    orgs = report.get_organizations()

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start", date(2000, 1, 1))
    with col2:
        end_date = st.date_input("End", date.today())

    selected_orgs = st.multiselect("Organization", orgs, placeholder="All")

    start = datetime(start_date.year, start_date.month, start_date.day)
    end = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
    org_filter = selected_orgs if selected_orgs else None

    r1, r2, r3 = st.columns(3)
    with r1:
        st.metric(
            "Registered", report.total_registered_participants(start, end, org_filter)
        )
    with r2:
        st.metric(
            "Attended", report.total_attended_participants(start, end, org_filter)
        )
    with r3:
        h = report.total_training_hours(start, end, org_filter)
        st.metric("Training Hours", f"{h:.0f}")

    r4, r5, r6 = st.columns(3)
    with r4:
        st.metric(
            "Individuals", report.total_individual_participants(start, end, org_filter)
        )
    with r5:
        g = report.government_percentage(start, end, org_filter)
        st.metric("Government", f"{g}%")
    with r6:
        st.metric("Courses", report.total_courses_conducted(start, end))

    st.subheader("Gender Distribution")
    gd = report.gender_distribution(start, end, org_filter)
    kind, show_legend, show_labels = _chart_controls("gender")
    _render_chart(gd, kind, show_legend, show_labels)

    st.subheader("Course Level Distribution")
    lv = report.course_level_distribution(start, end)
    kind, show_legend, show_labels = _chart_controls("level")
    _render_chart(lv, kind, show_legend, show_labels)

    if selected_orgs:
        st.subheader("Participants by Organization")
        od = report.participants_by_organization(start, end, selected_orgs)
        if od:
            st.dataframe(
                {"Organization": list(od.keys()), "Participants": list(od.values())},
                width="stretch",
            )

err_file = Path(ERR_XLSX)
if err_file.exists():
    err_rows = sum(ws.max_row - 1 for ws in load_workbook(err_file).worksheets)
    if err_rows:
        st.subheader("Normalization Errors")
        st.download_button(
            "Download error rows",
            err_file.read_bytes(),
            file_name=err_file.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
