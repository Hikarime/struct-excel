import logging
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from normalization import normalize_sheet

SRC_XLSX = "../data/dummy.xlsx"
TARGET_XLSX = "../dist/result.xlsx"
ERR_XLSX = "../dist/err.xlsx"

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

wb = load_workbook(SRC_XLSX)
err_wb = Workbook()

for sheet in wb:
    err_ws = err_wb.create_sheet(sheet.title)
    normalize_sheet(sheet, err_ws)

wb.save(TARGET_XLSX)
err_wb.save(ERR_XLSX)
